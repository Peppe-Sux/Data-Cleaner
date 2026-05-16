import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

"""
DATA CLEANING PIPELINE - PROFESSIONAL VERSION
--------------------------------------------
Demonstrated Skills:
1. Universal Data Loading: Dynamic file ingestion handling both CSV and Excel formats.
2. Robust String Sanitization: Trim operations, Title Case transformation, and character stripping.
3. Defensive Datetime Parsing: Mixed-format date normalization with day-first fallbacks.
4. Advanced Regex Filtering: Strategic row-masking via pattern matching (Emails/Phones).
5. OpenXML Presentation Styling: Aesthetic automated column resizing and header branding.
"""

def process_file(input_file, output_file=None, force_excel=False):
    """
    Cleans structural anomalies from CSV/Excel sources and saves standardized results.
    - If output_file is None, defaults to [name]_cleaned.[extension].
    - If force_excel=True, converts/saves the output as an executive-ready Excel file.
    """
    # Automatic output naming layers
    if output_file is None:
        base_name = input_file.rsplit('.', 1)[0]
        ext = '.xlsx' if force_excel else f".{input_file.rsplit('.', 1)[1]}"
        output_file = f"{base_name}_cleaned{ext}"

    # 1. Universal Data Loading
    print(f"[LOADING] Accessing data source: {input_file}...")
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    else:
        df = pd.read_excel(input_file)
    
    print(f" -> Initial records detected: {len(df)}")
    
    # --- 2. ADVANCED DATA CLEANING ENGINE ---
    
    # Step 1: Broad multi-column whitespace trim
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    
    # Step 2: Proper Case enforcement for customer names
    if 'First_Name' in df.columns:
        df['First_Name'] = df['First_Name'].str.title()
    
    # Step 3: Defensive Datetime Normalization (Fixes format variance bugs)
    if 'Registration_Date' in df.columns:
        # 'mixed' infers varying structures; 'dayfirst=True' ensures international layout handling
        df['Registration_Date'] = pd.to_datetime(
            df['Registration_Date'], 
            format='mixed', 
            dayfirst=True, 
            errors='coerce'
        ).dt.strftime('%Y-%m-%d')
    
    # Step 4: Strict Email Structural Validation via Regex Masking
    if 'Email' in df.columns:
        regex_email = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        mask_email = df['Email'].str.contains(regex_email, na=False, regex=True)
        print(f" -> Invalid email formats dropped: {len(df) - mask_email.sum()}")
        df = df[mask_email]
    
    # Step 5: Phone Number Sanitization (Strips out special chars, spaces, dashes)
    if 'Phone' in df.columns:
        df['Phone'] = df['Phone'].replace(r'\D', '', regex=True)
    
    # Step 6: Entity Deduplication based on unique keys
    if 'Email' in df.columns:
        initial_rows = len(df)
        df = df.drop_duplicates(subset=['Email'])
        print(f" -> Duplicate records purged: {initial_rows - len(df)}")
    
    # Step 7: Missing Value (Null) Allocation Management
    if 'City' in df.columns:
        df['City'] = df['City'].fillna('N/A')
    if 'Budget' in df.columns:
        df['Budget'] = df['Budget'].fillna(0)
    
    print(f"[CLEANING COMPLETE] Final consolidated dataset rows: {len(df)}")

    # --- 3. RE-RENDERING & EXTENSION ROUTING ---
    if force_excel or output_file.endswith('.xlsx'):
        # Save to native OpenXML Excel layout
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Apply professional corporate stylesheet themes
        wb = load_workbook(output_file)
        ws = wb.active
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Format Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        # Auto-fit column widths based on dynamic text metadata lengths
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 3
        
        wb.save(output_file)
        print(f"[SUCCESS] Exported styled spreadsheet: {output_file}")
        
    else:
        # Standard Raw Format saving
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        print(f"[SUCCESS] Exported raw flat file: {output_file}")


# --- DYNAMIC USE CASE EXECUTION ENGINE ---
if __name__ == "__main__":
    print("=== STARTING DEMO DATA PROCESSING PIPELINE ===\n")

    # Dynamically locate the directory where this script is living
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    
    # Generate dynamic absolute paths for source test files
    csv_test_path = os.path.join(SCRIPT_DIR, 'dirty_data.csv')
    xlsx_test_path = os.path.join(SCRIPT_DIR, 'dirty_data.xlsx')

    # Isolation wrapper to handle individual target test cases cleanly
    def safe_run_case(case_number, description, func, *args, **kwargs):
        print(f"\n--- CASE {case_number}: {description} ---")
        try:
            func(*args, **kwargs)
        except Exception as e:
            print(f"[CASE {case_number} ERROR] Could not complete execution block: {e}")

    # =========================================================================
    # 1. CSV -> CSV (Automatic name: 'dirty_data_cleaned.csv')
    # =========================================================================
    safe_run_case(1, "Standard pipeline from CSV to CSV with auto-naming rules", 
                  process_file, csv_test_path)

    # =========================================================================
    # 2. CSV -> CSV (Custom name: 'final_roster.csv')
    # =========================================================================
    custom_csv_out = os.path.join(SCRIPT_DIR, 'final_roster.csv')
    safe_run_case(2, "Pipeline execution from CSV to CSV with custom targeted output file", 
                  process_file, csv_test_path, output_file=custom_csv_out)

    # =========================================================================
    # 3. CSV -> EXCEL (Forced Conversion, Automatic name: 'dirty_data.xlsx')
    # =========================================================================
    safe_run_case(3, "Forced structure format conversion from flat CSV to styled Excel sheet", 
                  process_file, csv_test_path, force_excel=True)

    # =========================================================================
    # 4. CSV -> EXCEL (Forced Conversion with custom targeted filename)
    # =========================================================================
    custom_xlsx_out = os.path.join(SCRIPT_DIR, 'corporate_report.xlsx')
    safe_run_case(4, "Conversion execution from CSV to custom stylized executive Excel report", 
                  process_file, csv_test_path, output_file=custom_xlsx_out, force_excel=True)

    # =========================================================================
    # 5. EXCEL -> EXCEL (Automatic name: 'dirty_data_cleaned.xlsx')
    # =========================================================================
    safe_run_case(5, "Native processing from Excel to Excel using default naming layers", 
                  process_file, xlsx_test_path)

    # =========================================================================
    # 6. EXCEL -> EXCEL (Custom name: 'final_data.xlsx')
    # =========================================================================
    custom_xlsx_from_xlsx = os.path.join(SCRIPT_DIR, 'final_data.xlsx')
    safe_run_case(6, "Native execution from Excel source straight to custom target Excel file", 
                  process_file, xlsx_test_path, output_file=custom_xlsx_from_xlsx)

    # =========================================================================
    # 7. EXCEL -> CSV (Forced Extension Format Downgrade)
    # =========================================================================
    converted_csv_out = os.path.join(SCRIPT_DIR, 'converted_data.csv')
    safe_run_case(7, "Forced down-sampling export routing from rich Excel spreadsheet to flat CSV file", 
                  process_file, xlsx_test_path, output_file=converted_csv_out)
